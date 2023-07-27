import tkinter as tk
from tkinter import *
from openpyxl import * 
from tkinter import messagebox


signupfile = load_workbook('C:\\Users\\cabemaiwaie\\Desktop\\flashcards.xlsx')
sheet = signupfile.active


class FlashcardsApp(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        width = 350
        height = 500

        self.minsize(width, height)
        self.title('Flashcard Application')

        # creating container
        container = tk.Frame(self, width=350, height=500)
        container.pack(side="top", fill="both", expand=True)

        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.frames = {}

        for F in (Home, CreateFlashcards, FlashcardPack):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")
            self.show_frame(Home)

    #  Show window for the given page name
    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

    signupfile = load_workbook('C:\\Users\\cabemaiwaie\\Desktop\\flashcards.xlsx')
    sheet = signupfile.active
    
    def excel():
            # resize the width of columns in excel spreadsheet
            sheet.column_dimensions['A'].width = 30
            sheet.column_dimensions['B'].width = 20
            sheet.column_dimensions['C'].width = 30
            sheet.column_dimensions['D'].width = 30
            sheet.column_dimensions['E'].width = 20
            sheet.column_dimensions['F'].width = 40
            sheet.column_dimensions['G'].width = 50

            # add headings to excel spreadsheet
            sheet.cell(row=1, column=1).value = "Topic"
            sheet.cell(row=1, column=2).value = "Flashcard no."
            sheet.cell(row=1, column=3).value = "Front side"
            sheet.cell(row=1, column=4).value = "Back side"

            signupfile.save('C:\\Users\\cabemaiwaie\\Desktop\\flashcards.xlsx')

    excel()
    

class Home(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.parent = parent

        home_label = tk.Label(self, text="Home")
        home_label.pack(padx=10, pady=10)

        # Buttons
        flashcard_pack_btn = tk.Button(self, text="Create Flashcard Packs", bg='white', width=20,
                                height=10, font=('MS Sans Serif', 8),
                                command=lambda: controller.show_frame(FlashcardPack))
        create_flashcards_btn = tk.Button(self, text="Create Flashcards", bg='white', width=20, height=10, font=('MS Sans Serif', 8),
                              command=lambda: controller.show_frame(CreateFlashcards))
        flashcard_pack_btn.pack(padx=5, pady=10)
        create_flashcards_btn.pack(padx=5, pady=10)


class FlashcardPack(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        pack_label = tk.Label(self, text="Create Your Flashcard Packs")
        pack_label.pack(padx=10, pady=10)

        subject_label = tk.Label(self, text='Please enter the topic you would like for your flashcard pack:')
        subject_label.pack(padx=10, pady=10)
        subject_field = tk.Entry(self)
        subject_field.pack(padx=10, pady=10)

        def focus(event):
            subject_field.focuse_set()

        # function clears text entry boxes when called
        def clear():

            subject_field.delete(0, END)
           
        # function takes input from sign up form and writes it into the excel file
        def insert():

            # checking if entry is empty 
            if subject_field.get() == "":
                 messagebox.showinfo("No Entry", "Please enter a topic")
            else:
                # assigning the max row value up to which data is written in an excel sheet to variables
                current_row = sheet.max_row

                # get method returns user input as string which is written into excel spredsheet at a certain location
                sheet.cell(row=current_row + 1, column=1).value = subject_field.get()
                
                
                signupfile.save('C:\\Users\\cabemaiwaie\\Desktop\\flashcards.xlsx')
                
                subject_field.focus_set()

                clear()

        

        # bind method used to call the next focus function when enter key is pressed 
        subject_field.bind("<Return>", focus)
        
        submit_btn = tk.Button(self, text='Submit', command=insert)
        submit_btn.pack(padx=10, pady=10)



        

        
        
        


class CreateFlashcards(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)


if __name__ == "__main__":
    app = FlashcardsApp()
    app.mainloop()
